from config import base_path, cell_config, cases_sheet_name, variables_sheet_name
import os
import openpyxl
import json


class FileTool(object):
    # 1. 初始化
    def __init__(self, filename):
        # 1. 动态文件路径
        self.filename = base_path + os.sep + "data" + os.sep + filename
        print("要的打开的文件为：", self.filename)
        # 2. 打开文件 获取workbook对象
        self.workbook = openpyxl.load_workbook(self.filename)
        # 3. 获取sheet表单对象
        self.sheet = self.workbook[cases_sheet_name]
        # 4. 获取总行数 (读取数据，遍历数据使用)
        self.row = self.sheet.max_row
        print("总行数数据为：", self.row)
        # 5. 获取全局变量表
        self.variables_sheet = self.workbook[variables_sheet_name]

        # 6. 定义全局变量variables，存放解析后的变量字典
        self.variables = dict()
        self.load_variables()


    # 2. 读取Excel
    def read_excel(self):
        # 1. 新建空列表 （存储每行数据）
        case = list()
        # 2. 遍历每行数据
        for i in range(2, self.row + 1):
            # 新建空字典 （存储每行数据）
            data = dict()
            # 判断是否执行
            print(self.sheet.cell(i, 1).value)
            print(self.sheet.cell(i, cell_config.get("is_run")).value)
            if self.sheet.cell(i, cell_config.get("is_run")).value == "是":
                try:
                    # 读取数据 追加到字典
                    data['protocol'] = self.sheet.cell(i, cell_config.get("protocol")).value
                    data['path'] = self.sheet.cell(i, cell_config.get("path")).value
                    data['method'] = str(self.sheet.cell(i, cell_config.get("method")).value).lower()

                    headers = self.sheet.cell(i, cell_config.get("headers")).value
                    data['headers'] = eval(self.get_filled_variable(headers))
                    
                    params = self.sheet.cell(i, cell_config.get("params")).value
                    data['params'] = eval(self.get_filled_variable(params))
                    data['param_type'] = self.sheet.cell(i, cell_config.get("param_type")).value


                    expect = self.sheet.cell(i, cell_config.get("expect")).value
                    data['expect'] = eval(self.get_filled_variable(expect))

                    data['export_data'] = self.sheet.cell(i, cell_config.get("export_data")).value

                    # 记录每行数据的 行与列，执行完写入结果使用；
                    data['x_y'] = [i, cell_config.get("result")]
                    # 将字典追加到列表中
                    case.append(data)
                    # 将读取结果写入excel中
                    self.write_excel([i, cell_config.get("desc")], "数据读取完成～！")
                except Exception as e:
                    self.write_excel([i, cell_config.get("desc")], e)
        # 3. 将列表数据写入json
        self.write_json(case, "case.json")
        print("读取的数据为：", case)

    # 3. 写入Excel
    def write_excel(self, x_y, msg):
        try:
            # x_y参数的格式为列表 如: [2,5]
            self.sheet.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            # x_y参数的格式为列表 如: [2,5]
            self.sheet.cell(x_y[0], x_y[1]).value = e
        finally:
            # 保存excel
            self.workbook.save(self.filename)

    # 4. 读取Json
    def read_json(self, file_name="case.json"):
        file_name = base_path + os.sep + "data" + os.sep + file_name
        with open(file_name, "r", encoding="utf-8") as f:
            return json.load(f)

    # 5. 写入Json
    def write_json(self, case, file_name):
        file_name = base_path + os.sep + "data" + os.sep + file_name
        with open(file_name, "w", encoding="utf-8")as f:
            json.dump(case, f, indent=4, ensure_ascii=False)

    def get_filled_variable(self, text):
        result_text = text
        for key in self.variables:
            result_text = result_text.replace(key, self.variables[key])
        return result_text

    def load_variables(self):
        rows = self.variables_sheet.max_row + 1
        for i in range(2, rows):
            self.variables["{{"+self.variables_sheet.cell(i, 1).value+"}}"] = self.variables_sheet.cell(i, 2).value


if __name__ == '__main__':
    FileTool("测试用例_deom.xlsx").read_excel()
    # print(FileTool("Case01.xlsx").read_json())
